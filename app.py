from __future__ import annotations
# ════════════════════════════════════════════════════════════════════════
#  SenseOptics Метролог — Replit/mobile friendly версия
#  v5: пакетный режим ROI с отложенным расчётом (batch ROI / lazy metrics).
#
#  Главная идея версии:
#    • клик НИЧЕГО не считает — он только накапливает точки/ROI;
#    • метрики, overlay и экспорт строятся ТОЛЬКО по кнопке «Посчитать метрики»;
#    • маска строится ТОЛЬКО по своей кнопке;
#    • тяжёлые операции (Excel, JSON, PNG, ZIP, resize, mask) не выполняются
#      после каждого rerun — это и даёт ускорение на бесплатном Replit.
# ════════════════════════════════════════════════════════════════════════
import hashlib
import importlib.util
import io
import json
import math
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import cv2
import numpy as np
import streamlit as st
from PIL import Image as PILImage, ImageOps

try:
    from streamlit_image_coordinates import streamlit_image_coordinates as st_imgcoords
    HAS_CLICK = True
except Exception:
    st_imgcoords = None
    HAS_CLICK = False

# openpyxl импортируется ЛЕНИВО внутри build_xlsx_from_metrics().
# Здесь только проверяем наличие пакета, не загружая его.
HAS_XLSX = importlib.util.find_spec("openpyxl") is not None

st.set_page_config(
    page_title="SenseOptics Метролог",
    page_icon="📐",
    layout="wide",
    initial_sidebar_state="collapsed",
)

APP_VERSION = "app334-mobile-hotfix-v5-batch-roi-replit-free-2026-06-24"

# Режимы (внутренние константы, чтобы не сравнивать строки по всему коду)
MODE_LINE = "Линия"
MODE_CIRCLE = "Окружность (3 точки)"
MODE_POLY = "Площадь (полигон)"
MODE_INCL = "Включения"
MODE_MASK = "ROI маски"
MODE_CALIB = "Калибровка"
MODES = [MODE_LINE, MODE_CIRCLE, MODE_POLY, MODE_INCL, MODE_MASK, MODE_CALIB]

GREEN = (40, 180, 70)
BLUE = (40, 120, 230)
ORANGE = (255, 150, 0)
RED = (220, 50, 50)
PURPLE = (165, 80, 205)
BLACK = (20, 20, 20)


# ════════════════════════════════════════════════════════════════════════
#  Состояние
# ════════════════════════════════════════════════════════════════════════
def _init_state() -> None:
    defaults = {
        # --- пакетный ROI ---
        "draft_points": [],        # точки текущего незавершённого ROI
        "draft_rois": [],          # накопленные ROI (ещё не обязательно рассчитаны)
        "measurements": [],        # рассчитанные метрики (= cached_metrics["measurements"])
        "metrics_dirty": False,    # True, если ROI менялись и нужен пересчёт
        "cached_metrics": None,    # последний результат calculate_all_metrics
        "cached_overlay": None,    # overlay (RGB) после расчёта
        "cached_exports": None,    # кэш лёгких экспортов после расчёта
        "current_image_id": None,  # идентификатор текущего изображения
        "last_uploaded_path": None,  # путь к последнему сохранённому файлу в uploads
        "roi_counter": 0,          # для уникальных id ROI

        # --- калибровка ---
        "px_per_mm": None,
        "calibration_info": None,

        # --- клики/режим/полноэкранный ---
        "last_click_id": None,
        "fullscreen_mode": False,
        "canvas_display_width": 420,
        "inclusion_fast_mode": True,
        "mask_roi_shape": "polygon",
        "show_control_preview_v5": False,

        # --- маска ---
        "mask": None,
        "mask_overlay": None,
        "mask_stats": None,
        "mask_settings": None,

        # --- надёжное хранение изображения (file_uploader на телефоне/Replit
        #     иногда теряет файл при rerun) ---
        "stored_image_bytes": None,
        "stored_image_name": None,
        "stored_image_id": None,
        "stored_image_source": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_mask() -> None:
    st.session_state.mask = None
    st.session_state.mask_overlay = None
    st.session_state.mask_stats = None
    st.session_state.mask_settings = None


def clear_work_state(reset_calibration: bool = True) -> None:
    """Полная очистка рабочего состояния (при смене изображения/ширины)."""
    st.session_state.draft_points = []
    st.session_state.draft_rois = []
    st.session_state.measurements = []
    st.session_state.cached_metrics = None
    st.session_state.cached_overlay = None
    st.session_state.cached_exports = None
    st.session_state.metrics_dirty = False
    st.session_state.last_click_id = None
    st.session_state.roi_counter = 0
    reset_mask()
    if reset_calibration:
        st.session_state.px_per_mm = None
        st.session_state.calibration_info = None


_init_state()


# ════════════════════════════════════════════════════════════════════════
#  CSS / мобильный вид
# ════════════════════════════════════════════════════════════════════════
st.markdown(
    """
<style>
    /* Полностью прячем sidebar, чтобы мобильный Replit не открывал левую панель */
    section[data-testid="stSidebar"] {display: none !important;}
    [data-testid="collapsedControl"] {display: none !important;}

    /* Главное: не даём странице уезжать вправо на телефоне */
    html, body, .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"] {
        overflow-x: hidden !important;
        max-width: 100vw !important;
    }
    .block-container {
        max-width: 100% !important;
        padding-top: 0.75rem;
        padding-left: clamp(0.35rem, 1.6vw, 1.2rem);
        padding-right: clamp(0.35rem, 1.6vw, 1.2rem);
    }

    h1 {
        font-size: clamp(1.35rem, 7vw, 2.2rem) !important;
        line-height: 1.08 !important;
        overflow-wrap: anywhere !important;
        word-break: break-word !important;
    }
    h2, h3, p, span, label, div {
        overflow-wrap: anywhere;
    }

    img, canvas, iframe {
        max-width: 100% !important;
    }
    div[data-testid="stImage"] img {
        max-width: 100% !important;
        height: auto !important;
    }

    div[data-testid="stFileUploader"] section {
        border-radius: 16px;
        min-height: 86px;
        max-width: 100% !important;
    }
    div[data-testid="stFileUploader"] section * {
        max-width: 100% !important;
        overflow-wrap: anywhere !important;
    }

    div.stButton > button, div[data-testid="stDownloadButton"] > button {
        min-height: 44px;
        border-radius: 10px;
        width: 100%;
        white-space: normal !important;
    }
    div[role="radiogroup"] label {
        padding: 0.25rem 0.25rem;
        min-height: 36px;
    }

    .so-hint {
        padding: 0.65rem 0.8rem;
        border-radius: 12px;
        background: rgba(128, 128, 128, 0.08);
        margin: 0.4rem 0 0.7rem 0;
        font-size: 0.92rem;
    }
    .so-small {
        opacity: 0.78;
        font-size: 0.84rem;
        line-height: 1.35;
    }
    .so-status {
        padding: 0.5rem 0.75rem;
        border-radius: 10px;
        background: rgba(40, 120, 230, 0.10);
        font-size: 0.9rem;
        margin: 0.25rem 0 0.5rem 0;
    }
    .so-dirty {
        background: rgba(255, 150, 0, 0.16) !important;
    }
    .so-version {
        opacity: 0.55;
        font-size: 0.78rem;
        margin-top: -0.35rem;
        margin-bottom: 0.45rem;
    }

    @media (max-width: 760px) {
        .block-container {
            padding-left: 0.28rem;
            padding-right: 0.28rem;
            padding-top: 0.45rem;
        }
        h1 {font-size: 1.32rem !important;}
        h2 {font-size: 1.10rem !important;}
        h3 {font-size: 1.00rem !important;}
        div.stButton > button, div[data-testid="stDownloadButton"] > button {
            min-height: 46px;
            font-size: 0.88rem;
        }
        /* Streamlit columns иногда не успевают схлопнуться внутри Replit WebView.
           Это заставляет колонки идти одной под другой и убирает горизонтальный скролл. */
        div[data-testid="stHorizontalBlock"] {
            flex-wrap: wrap !important;
        }
        div[data-testid="stHorizontalBlock"] > div {
            min-width: min(100%, 320px) !important;
            flex: 1 1 100% !important;
        }
    }
</style>
""",
    unsafe_allow_html=True,
)


# ════════════════════════════════════════════════════════════════════════
#  Геометрия (чистые функции, без Streamlit)
# ════════════════════════════════════════════════════════════════════════
def dist(a, b) -> float:
    return math.hypot(b[0] - a[0], b[1] - a[1])


def poly_area(pts) -> float:
    if len(pts) < 3:
        return 0.0
    s = 0.0
    for i in range(len(pts)):
        x1, y1 = pts[i]
        x2, y2 = pts[(i + 1) % len(pts)]
        s += x1 * y2 - x2 * y1
    return abs(s) / 2.0


def poly_perimeter(pts) -> float:
    if len(pts) < 2:
        return 0.0
    return sum(dist(pts[i], pts[(i + 1) % len(pts)]) for i in range(len(pts)))


def circumcircle(p1, p2, p3) -> Tuple[Optional[Tuple[float, float]], Optional[float]]:
    """Окружность по трём точкам. Возвращает (center, radius) или (None, None) для коллинеарных."""
    ax, ay = float(p1[0]), float(p1[1])
    bx, by = float(p2[0]), float(p2[1])
    cx, cy = float(p3[0]), float(p3[1])
    d = 2.0 * (ax * (by - cy) + bx * (cy - ay) + cx * (ay - by))
    if abs(d) < 1e-9:
        return None, None
    a2 = ax * ax + ay * ay
    b2 = bx * bx + by * by
    c2 = cx * cx + cy * cy
    ux = (a2 * (by - cy) + b2 * (cy - ay) + c2 * (ay - by)) / d
    uy = (a2 * (cx - bx) + b2 * (ax - cx) + c2 * (bx - ax)) / d
    r = math.hypot(ax - ux, ay - uy)
    return (ux, uy), r


# ════════════════════════════════════════════════════════════════════════
#  Расчёт всех метрик (ТОЛЬКО по кнопке)
# ════════════════════════════════════════════════════════════════════════
def calculate_all_metrics(
    draft_rois: List[Dict[str, Any]],
    px_per_mm: Optional[float],
    image_shape: Tuple[int, int],
    mask: Optional[np.ndarray] = None,
) -> Dict[str, Any]:
    """Единая точка расчёта. Считает метрики по всем накопленным ROI."""
    H, W = int(image_shape[0]), int(image_shape[1])
    measurements: List[Dict[str, Any]] = []
    warnings: List[str] = []
    n_inclusions = 0
    total_line_mm = 0.0
    analysis_area_px = 0.0  # суммарная площадь mask_roi/polygon для плотности включений

    for roi in draft_rois:
        t = roi.get("type")
        pts = roi.get("points", [])
        rec: Dict[str, Any] = {
            "id": roi.get("id"),
            "type": t,
            "label": roi.get("label", ""),
            "points": pts,
        }

        if t == "line":
            if len(pts) < 2:
                warnings.append(f"{roi.get('id')}: для линии нужно 2 точки.")
                continue
            d = dist(pts[0], pts[1])
            rec["length_px"] = round(d, 2)
            if px_per_mm:
                rec["length_mm"] = round(d / px_per_mm, 4)
                total_line_mm += rec["length_mm"]

        elif t == "circle_3pt":
            if len(pts) < 3:
                warnings.append(f"{roi.get('id')}: для окружности нужно 3 точки.")
                continue
            center, r = circumcircle(pts[0], pts[1], pts[2])
            if center is None:
                warnings.append(f"{roi.get('id')}: 3 точки на одной прямой — окружность не построена.")
                continue
            rec["center_x"] = round(center[0], 2)
            rec["center_y"] = round(center[1], 2)
            rec["radius_px"] = round(r, 2)
            rec["diameter_px"] = round(2 * r, 2)
            rec["circumference_px"] = round(2 * math.pi * r, 2)
            rec["area_px2"] = round(math.pi * r * r, 1)
            if px_per_mm:
                rec["radius_mm"] = round(r / px_per_mm, 4)
                rec["diameter_mm"] = round(2 * r / px_per_mm, 4)
                rec["circumference_mm"] = round(2 * math.pi * r / px_per_mm, 4)
                rec["area_mm2"] = round(math.pi * r * r / (px_per_mm * px_per_mm), 4)

        elif t == "polygon":
            if len(pts) < 3:
                warnings.append(f"{roi.get('id')}: для полигона нужно минимум 3 точки.")
                continue
            a = poly_area(pts)
            p = poly_perimeter(pts)
            analysis_area_px += a
            rec["area_px2"] = round(a, 1)
            rec["perimeter_px"] = round(p, 2)
            rec["n_vertices"] = len(pts)
            if px_per_mm:
                rec["area_mm2"] = round(a / (px_per_mm * px_per_mm), 4)
                rec["perimeter_mm"] = round(p / px_per_mm, 4)

        elif t == "inclusion_point":
            if not pts:
                continue
            n_inclusions += 1
            rec["x"] = int(pts[0][0])
            rec["y"] = int(pts[0][1])

        elif t == "mask_roi":
            if len(pts) < 3:
                warnings.append(f"{roi.get('id')}: для ROI маски нужно минимум 3 точки.")
                continue
            a = poly_area(pts)
            analysis_area_px += a
            rec["roi_area_px2"] = round(a, 1)
            if px_per_mm:
                rec["roi_area_mm2"] = round(a / (px_per_mm * px_per_mm), 4)
            if mask is not None and a > 0:
                poly_mask = np.zeros((H, W), np.uint8)
                cv2.fillPoly(poly_mask, [np.array(pts, np.int32)], 255)
                inside = int(np.count_nonzero((mask > 0) & (poly_mask > 0)))
                rec["mask_px_inside"] = inside
                rec["mask_percent_inside"] = round(100.0 * inside / a, 3)
                rec["mask_fraction_inside"] = round(inside / a, 5)
                if px_per_mm:
                    rec["mask_area_inside_mm2"] = round(inside / (px_per_mm * px_per_mm), 5)
        else:
            warnings.append(f"{roi.get('id')}: неизвестный тип ROI '{t}'.")
            continue

        measurements.append(rec)

    summary = {
        "n_rois": len(draft_rois),
        "n_lines": sum(1 for r in draft_rois if r.get("type") == "line"),
        "n_circles": sum(1 for r in draft_rois if r.get("type") == "circle_3pt"),
        "n_polygons": sum(1 for r in draft_rois if r.get("type") == "polygon"),
        "n_inclusions": n_inclusions,
        "n_mask_rois": sum(1 for r in draft_rois if r.get("type") == "mask_roi"),
        "px_per_mm": px_per_mm,
        "total_line_length_mm": round(total_line_mm, 4) if px_per_mm else None,
    }

    statistics: Dict[str, Any] = {}
    if n_inclusions and analysis_area_px > 0 and px_per_mm:
        area_mm2 = analysis_area_px / (px_per_mm * px_per_mm)
        statistics["analysis_area_mm2"] = round(area_mm2, 4)
        statistics["inclusion_density_per_mm2"] = (
            round(n_inclusions / area_mm2, 4) if area_mm2 > 0 else None
        )
    elif n_inclusions:
        statistics["note"] = (
            "Для плотности включений задайте калибровку и хотя бы один полигон / ROI области анализа."
        )

    return {
        "measurements": measurements,
        "summary": summary,
        "statistics": statistics,
        "warnings": warnings,
    }


# ════════════════════════════════════════════════════════════════════════
#  Отрисовка
# ════════════════════════════════════════════════════════════════════════
def _put(img, text, org, color) -> None:
    if not text:
        return
    cv2.putText(img, text, org, cv2.FONT_HERSHEY_SIMPLEX, 0.48, color, 1, cv2.LINE_AA)


def draw_workspace(
    work_bgr: np.ndarray,
    draft_rois: List[Dict[str, Any]],
    draft_points: List[List[int]],
    mode: str,
    measurements: Optional[List[Dict[str, Any]]] = None,
) -> np.ndarray:
    """
    Лёгкая отрисовка. Если measurements не передан — рисуем только фигуры и имена
    (без расчётов), что делает перерисовку после каждого клика дешёвой.
    Если measurements передан (уже посчитан и закэширован) — добавляем значения.
    Никаких пересчётов метрик здесь не происходит.
    """
    rgb = cv2.cvtColor(work_bgr, cv2.COLOR_BGR2RGB).copy()
    meas_by_id = {m.get("id"): m for m in (measurements or [])}
    incl_n = 0

    for roi in draft_rois:
        t = roi.get("type")
        pts = roi.get("points", [])
        m = meas_by_id.get(roi.get("id"))

        if t == "line" and len(pts) >= 2:
            p1, p2 = tuple(pts[0]), tuple(pts[1])
            cv2.line(rgb, p1, p2, GREEN, 2)
            cv2.circle(rgb, p1, 4, GREEN, -1)
            cv2.circle(rgb, p2, 4, GREEN, -1)
            label = roi.get("label", "")
            if m is not None:
                label = (
                    f"{m.get('length_mm')} мм" if m.get("length_mm") is not None
                    else f"{m.get('length_px')} px"
                )
            mx, my = (p1[0] + p2[0]) // 2, (p1[1] + p2[1]) // 2
            _put(rgb, label, (mx + 5, my - 7), GREEN)

        elif t == "circle_3pt" and len(pts) >= 3:
            center, r = circumcircle(pts[0], pts[1], pts[2])
            if center is not None:
                c = (int(round(center[0])), int(round(center[1])))
                ri = int(round(r))
                cv2.circle(rgb, c, ri, BLUE, 2)
                cv2.circle(rgb, c, 3, BLUE, -1)
                for p in pts:
                    cv2.circle(rgb, tuple(p), 4, BLUE, -1)
                label = roi.get("label", "")
                if m is not None:
                    label = (
                        f"D {m.get('diameter_mm')} мм" if m.get("diameter_mm") is not None
                        else f"D {m.get('diameter_px')} px"
                    )
                _put(rgb, label, (c[0] + 5, max(15, c[1] - ri - 7)), BLUE)
            else:
                for p in pts:
                    cv2.circle(rgb, tuple(p), 4, RED, 2)

        elif t in ("polygon", "mask_roi") and len(pts) >= 2:
            col = ORANGE if t == "polygon" else PURPLE
            arr = np.array(pts, np.int32)
            cv2.polylines(rgb, [arr], True, col, 2)
            for p in pts:
                cv2.circle(rgb, tuple(p), 4, col, -1)
            cx = int(np.mean([p[0] for p in pts]))
            cy = int(np.mean([p[1] for p in pts]))
            label = roi.get("label", "")
            if m is not None and t == "polygon":
                label = (
                    f"{m.get('area_mm2')} мм²" if m.get("area_mm2") is not None
                    else f"{m.get('area_px2')} px²"
                )
            _put(rgb, label, (cx + 5, cy), col)

        elif t == "inclusion_point" and pts:
            incl_n += 1
            x, y = int(pts[0][0]), int(pts[0][1])
            cv2.circle(rgb, (x, y), 8, RED, 2)
            _put(rgb, str(incl_n), (x + 10, y - 6), RED)

    # Черновые точки текущего ROI
    for p in draft_points:
        cv2.drawMarker(rgb, tuple(p), BLACK, cv2.MARKER_CROSS, 14, 2)
    if mode in (MODE_POLY, MODE_MASK) and len(draft_points) >= 2:
        cv2.polylines(rgb, [np.array(draft_points, np.int32)], False, ORANGE, 1)
    if mode == MODE_CALIB:
        for i, p in enumerate(draft_points):
            cv2.circle(rgb, tuple(p), 9, ORANGE, 2)
            _put(rgb, f"C{i + 1}", (p[0] + 10, p[1] - 6), ORANGE)
        if len(draft_points) == 2:
            cv2.line(rgb, tuple(draft_points[0]), tuple(draft_points[1]), ORANGE, 1)

    return rgb


# ════════════════════════════════════════════════════════════════════════
#  Загрузка / декодирование (кэшируется)
# ════════════════════════════════════════════════════════════════════════
def file_fingerprint(data: bytes, name: str) -> str:
    h = hashlib.sha1()
    h.update(name.encode("utf-8", errors="ignore"))
    h.update(str(len(data)).encode())
    h.update(data[:1024 * 1024])
    return h.hexdigest()[:14]


@st.cache_data(show_spinner=False)
def decode_image(data: bytes, name: str = "") -> Tuple[Optional[np.ndarray], Optional[str]]:
    """BGR OpenCV. Сначала OpenCV, затем Pillow fallback для проблемных TIFF/PNG."""
    arr = np.frombuffer(data, np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is not None:
        return img, None
    try:
        pil = PILImage.open(io.BytesIO(data))
        pil = ImageOps.exif_transpose(pil)
        pil = pil.convert("RGB")
        rgb = np.array(pil)
        return cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR), None
    except Exception as e:
        return None, f"Не удалось прочитать изображение {name!r}: {e}"


@st.cache_data(show_spinner=False)
def cached_downscale(img: np.ndarray, max_w: int) -> Tuple[np.ndarray, float]:
    h, w = img.shape[:2]
    if w <= max_w:
        return img.copy(), 1.0
    scale = max_w / float(w)
    new_size = (int(round(w * scale)), int(round(h * scale)))
    return cv2.resize(img, new_size, interpolation=cv2.INTER_AREA), scale


# ════════════════════════════════════════════════════════════════════════
#  Маска тёмных/светлых включений (строится ТОЛЬКО по кнопке)
# ════════════════════════════════════════════════════════════════════════
def _odd(v: int) -> int:
    v = int(v)
    return v if v % 2 == 1 else v + 1


def _mask_roi_region(work_bgr: np.ndarray, mask_rois: List[Dict[str, Any]]) -> Optional[np.ndarray]:
    """Бинарная область из mask_roi полигонов или None, если их нет."""
    if not mask_rois:
        return None
    h, w = work_bgr.shape[:2]
    region = np.zeros((h, w), np.uint8)
    drew = False
    for roi in mask_rois:
        pts = roi.get("points", [])
        if len(pts) >= 3:
            cv2.fillPoly(region, [np.array(pts, np.int32)], 255)
            drew = True
    return region if drew else None


def build_inclusion_mask(
    work_bgr: np.ndarray,
    threshold_mode: str = "Otsu",
    manual_threshold: int = 90,
    min_area: int = 4,
    max_area: int = 0,
    detect_bright: bool = False,
    blur_size: int = 3,
    close_iter: int = 1,
    region: Optional[np.ndarray] = None,
) -> np.ndarray:
    gray = cv2.cvtColor(work_bgr, cv2.COLOR_BGR2GRAY)
    if blur_size > 1:
        gray = cv2.GaussianBlur(gray, (_odd(blur_size), _odd(blur_size)), 0)

    if threshold_mode == "Otsu":
        flag = cv2.THRESH_BINARY if detect_bright else cv2.THRESH_BINARY_INV
        _, mask = cv2.threshold(gray, 0, 255, flag + cv2.THRESH_OTSU)
    else:
        if detect_bright:
            mask = (gray >= int(manual_threshold)).astype(np.uint8) * 255
        else:
            mask = (gray <= int(manual_threshold)).astype(np.uint8) * 255

    # Ограничиваем маску областью mask_roi, если она задана
    if region is not None:
        mask = cv2.bitwise_and(mask, region)

    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel, iterations=1)
    if close_iter > 0:
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=int(close_iter))

    n, lab, stats, _ = cv2.connectedComponentsWithStats(mask, connectivity=8)
    clean = np.zeros_like(mask)
    for label in range(1, n):
        area = int(stats[label, cv2.CC_STAT_AREA])
        if area < int(min_area):
            continue
        if max_area and area > int(max_area):
            continue
        clean[lab == label] = 255
    return clean


def make_mask_overlay(work_bgr: np.ndarray, mask: np.ndarray, alpha: float = 0.45) -> np.ndarray:
    rgb = cv2.cvtColor(work_bgr, cv2.COLOR_BGR2RGB).copy()
    color = np.zeros_like(rgb)
    color[:, :] = RED
    m = mask > 0
    rgb[m] = cv2.addWeighted(rgb, 1.0 - alpha, color, alpha, 0)[m]
    return rgb


def mask_stats(mask: np.ndarray, ppm: Optional[float]) -> Dict[str, Any]:
    n, _, stats, _ = cv2.connectedComponentsWithStats(mask, connectivity=8)
    areas = [int(stats[label, cv2.CC_STAT_AREA]) for label in range(1, n)]
    total_px = int(np.count_nonzero(mask > 0))
    out: Dict[str, Any] = {
        "n_objects": n - 1,
        "total_mask_px": total_px,
        "area_fraction": round(float(np.mean(mask > 0)), 6),
        "max_object_area_px": int(max(areas)) if areas else 0,
        "mean_object_area_px": round(float(np.mean(areas)), 2) if areas else 0.0,
    }
    if ppm:
        out["total_area_mm2"] = round(total_px / (ppm * ppm), 5)
        out["max_object_area_mm2"] = round(out["max_object_area_px"] / (ppm * ppm), 5)
        out["mean_object_area_mm2"] = round(out["mean_object_area_px"] / (ppm * ppm), 5)
    return out


def gray_png_bytes(gray: np.ndarray) -> bytes:
    ok, buf = cv2.imencode(".png", gray)
    return buf.tobytes() if ok else b""


def png_bytes(rgb) -> bytes:
    ok, buf = cv2.imencode(".png", cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR))
    return buf.tobytes() if ok else b""


# ════════════════════════════════════════════════════════════════════════
#  Надёжное хранение загруженного изображения
# ════════════════════════════════════════════════════════════════════════
UPLOAD_DIR = Path(__file__).parent / "uploads"
LATEST_JSON = UPLOAD_DIR / "latest_image.json"


def _safe_filename(name: str) -> str:
    name = Path(name or "image.png").name
    name = re.sub(r"[^A-Za-zА-Яа-я0-9._-]+", "_", name)
    return name[:120] or "image.png"


def save_last_image(data: bytes, name: str, file_id: str) -> Optional[str]:
    if not data:
        return None
    try:
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        safe = _safe_filename(name)
        ext = Path(safe).suffix or ".png"
        stem = Path(safe).stem or "image"
        out = UPLOAD_DIR / f"last_{file_id}_{stem}{ext}"
        if not out.exists() or out.stat().st_size != len(data):
            out.write_bytes(data)
        LATEST_JSON.write_text(
            json.dumps({"path": str(out), "name": name, "file_id": file_id}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return str(out)
    except Exception:
        return None


def load_last_image_from_disk() -> Tuple[Optional[bytes], Optional[str], Optional[str], Optional[str]]:
    try:
        if not LATEST_JSON.exists():
            return None, None, None, None
        meta = json.loads(LATEST_JSON.read_text(encoding="utf-8"))
        path = Path(meta.get("path", ""))
        if not path.exists() or not path.is_file():
            return None, None, None, None
        data = path.read_bytes()
        name = meta.get("name") or path.name
        file_id = meta.get("file_id") or file_fingerprint(data, name)
        return data, name, file_id, str(path)
    except Exception:
        return None, None, None, None


def remember_image(data: bytes, name: str, source: str = "upload") -> str:
    file_id = file_fingerprint(data, name)
    st.session_state.stored_image_bytes = data
    st.session_state.stored_image_name = name
    st.session_state.stored_image_id = file_id
    st.session_state.stored_image_source = source
    disk_path = save_last_image(data, name, file_id)
    if disk_path:
        st.session_state.stored_image_source = f"{source}+disk"
        st.session_state.last_uploaded_path = disk_path
    return file_id


# ════════════════════════════════════════════════════════════════════════
#  Экспорт (строится только после расчёта / по кнопке)
# ════════════════════════════════════════════════════════════════════════
class _NpEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, np.integer):
            return int(o)
        if isinstance(o, np.floating):
            return float(o)
        if isinstance(o, np.ndarray):
            return o.tolist()
        return super().default(o)


def safe_json(d) -> str:
    return json.dumps(d, ensure_ascii=False, indent=2, cls=_NpEncoder)


def safe_stem(name: str) -> str:
    return Path(name).stem.replace(" ", "_") or "image"


def build_project_json(
    src_name: str,
    work_shape: Tuple[int, int],
    work_width: int,
    ppm: Optional[float],
    calibration_info: Optional[Dict[str, Any]],
    draft_rois: List[Dict[str, Any]],
    metrics: Optional[Dict[str, Any]],
    mask_settings: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    """Лёгкий JSON проекта. Без изображения и без base64."""
    metrics = metrics or {}
    return {
        "version": APP_VERSION,
        "image_name": src_name,
        "image_shape": [int(work_shape[0]), int(work_shape[1])],
        "work_width": int(work_width),
        "px_per_mm": ppm,
        "calibration_info": calibration_info,
        "draft_rois": draft_rois,
        "measurements": metrics.get("measurements", []),
        "mask_parameters": mask_settings,
        "summary": metrics.get("summary", {}),
        "statistics": metrics.get("statistics", {}),
    }


CSV_KEYS = [
    "id", "label", "type",
    "length_px", "length_mm",
    "diameter_px", "diameter_mm", "radius_px", "radius_mm",
    "circumference_px", "circumference_mm",
    "perimeter_px", "perimeter_mm",
    "area_px2", "area_mm2",
    "center_x", "center_y", "x", "y",
    "roi_area_px2", "roi_area_mm2", "mask_px_inside", "mask_percent_inside",
]


def build_csv_from_metrics(metrics: Optional[Dict[str, Any]]) -> str:
    metrics = metrics or {}
    lines = [",".join(CSV_KEYS)]
    for m in metrics.get("measurements", []):
        lines.append(",".join(str(m.get(k, "")) for k in CSV_KEYS))
    lines.append("")
    for k, v in (metrics.get("summary") or {}).items():
        lines.append(f"summary_{k},{v}")
    for k, v in (metrics.get("statistics") or {}).items():
        lines.append(f"stat_{k},{v}")
    return "\n".join(lines)


def build_xlsx_from_metrics(metrics: Optional[Dict[str, Any]], project: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook  # ленивый импорт

    metrics = metrics or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Замеры"
    ws.append(CSV_KEYS)
    for m in metrics.get("measurements", []):
        ws.append([m.get(k) for k in CSV_KEYS])

    ws_sum = wb.create_sheet("Сводка")
    for k, v in (metrics.get("summary") or {}).items():
        ws_sum.append([k, v])

    ws_stat = wb.create_sheet("Статистика")
    for k, v in (metrics.get("statistics") or {}).items():
        ws_stat.append([k, v])

    ws_info = wb.create_sheet("Инфо")
    for k in ("version", "image_name", "work_width", "px_per_mm"):
        ws_info.append([k, project.get(k)])
    if metrics.get("warnings"):
        ws_info.append([])
        ws_info.append(["warnings"])
        for w in metrics["warnings"]:
            ws_info.append([w])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_project_zip(
    image_data: bytes,
    src_name: str,
    project: Dict[str, Any],
    metrics: Optional[Dict[str, Any]],
    overlay_rgb: Optional[np.ndarray],
    mask: Optional[np.ndarray],
    mask_overlay_rgb: Optional[np.ndarray],
) -> bytes:
    """ZIP проекта. Изображение кладём как файл (без base64)."""
    buf = io.BytesIO()
    ext = (Path(src_name).suffix or ".png").lower()
    if ext not in (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"):
        ext = ".png"
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        if image_data:
            z.writestr(f"image{ext}", image_data)
        z.writestr("metrology_project.json", safe_json(project))
        if HAS_XLSX and metrics:
            z.writestr("measurements.xlsx", build_xlsx_from_metrics(metrics, project))
        else:
            z.writestr("measurements.csv", build_csv_from_metrics(metrics))
        if overlay_rgb is not None:
            z.writestr("overlay.png", png_bytes(overlay_rgb))
        if mask is not None:
            z.writestr("mask.png", gray_png_bytes(mask))
        if mask_overlay_rgb is not None:
            z.writestr("mask_overlay.png", png_bytes(mask_overlay_rgb))
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════
#  ROI / клики (НИЧЕГО не считают)
# ════════════════════════════════════════════════════════════════════════
_TYPE_NAMES = {
    "line": "Line",
    "circle_3pt": "Circle",
    "polygon": "Polygon",
    "inclusion_point": "Inclusion",
    "mask_roi": "Mask ROI",
}


def _type_label(roi_type: str) -> str:
    base = _TYPE_NAMES.get(roi_type, roi_type)
    n = sum(1 for r in st.session_state.draft_rois if r.get("type") == roi_type) + 1
    return f"{base} {n}"


def make_roi(roi_type: str, points, closed: Optional[bool] = None, label: Optional[str] = None) -> Dict[str, Any]:
    st.session_state.roi_counter += 1
    rid = f"roi_{st.session_state.roi_counter:04d}"
    roi: Dict[str, Any] = {
        "id": rid,
        "type": roi_type,
        "points": [[int(round(p[0])), int(round(p[1]))] for p in points],
        "label": label or _type_label(roi_type),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    if closed is not None:
        roi["closed"] = bool(closed)
    return roi


def add_roi(roi: Dict[str, Any]) -> None:
    st.session_state.draft_rois.append(roi)
    st.session_state.metrics_dirty = True


def clamp_click(x: int, y: int, W: int, H: int) -> Tuple[int, int]:
    return max(0, min(W - 1, x)), max(0, min(H - 1, y))


def handle_click(x: int, y: int, mode: str) -> None:
    """Только накопление точек/ROI. Никаких метрик."""
    if mode == MODE_INCL and st.session_state.inclusion_fast_mode:
        add_roi(make_roi("inclusion_point", [[x, y]]))
        return
    if mode == MODE_CALIB:
        if len(st.session_state.draft_points) >= 2:
            st.session_state.draft_points = [[x, y]]
        else:
            st.session_state.draft_points.append([x, y])
        return
    st.session_state.draft_points.append([x, y])


def commit_draft(mode: str, mask_shape: str) -> Tuple[bool, Optional[str]]:
    """Превращает накопленные draft_points в ROI. Возвращает (ok, error)."""
    pts = st.session_state.draft_points
    if mode == MODE_LINE:
        if len(pts) != 2:
            return False, "Для линии нужно ровно 2 точки."
        add_roi(make_roi("line", pts))
    elif mode == MODE_CIRCLE:
        if len(pts) != 3:
            return False, "Для окружности нужно ровно 3 точки."
        add_roi(make_roi("circle_3pt", pts))
    elif mode == MODE_POLY:
        if len(pts) < 3:
            return False, "Для полигона нужно минимум 3 точки."
        add_roi(make_roi("polygon", pts, closed=True))
    elif mode == MODE_INCL:
        if len(pts) < 1:
            return False, "Поставьте хотя бы одну точку включения."
        for p in pts:
            add_roi(make_roi("inclusion_point", [p]))
    elif mode == MODE_MASK:
        if mask_shape == "rect":
            if len(pts) != 2:
                return False, "Для прямоугольной ROI нужно 2 точки (противоположные углы)."
            (x1, y1), (x2, y2) = pts
            rect = [[x1, y1], [x2, y1], [x2, y2], [x1, y2]]
            add_roi(make_roi("mask_roi", rect, closed=True))
        else:
            if len(pts) < 3:
                return False, "Для полигональной ROI маски нужно минимум 3 точки."
            add_roi(make_roi("mask_roi", pts, closed=True))
    elif mode == MODE_CALIB:
        return False, "Калибровка применяется кнопкой «Применить калибровку», а не как ROI."
    st.session_state.draft_points = []
    return True, None


def undo_last_draft_point() -> bool:
    if st.session_state.draft_points:
        st.session_state.draft_points.pop()
        return True
    return False


def delete_last_roi() -> bool:
    if st.session_state.draft_rois:
        st.session_state.draft_rois.pop()
        st.session_state.metrics_dirty = True
        return True
    return False


def run_calculation(work_bgr: np.ndarray, mode: str) -> None:
    """Единственное место, где считаются метрики, overlay и лёгкие экспорты."""
    H, W = work_bgr.shape[:2]
    metrics = calculate_all_metrics(
        st.session_state.draft_rois,
        st.session_state.px_per_mm,
        (H, W),
        st.session_state.mask,
    )
    overlay_calc = draw_workspace(work_bgr, st.session_state.draft_rois, [], mode, metrics["measurements"])
    project = build_project_json(
        st.session_state.stored_image_name or "image",
        (H, W),
        W,
        st.session_state.px_per_mm,
        st.session_state.calibration_info,
        st.session_state.draft_rois,
        metrics,
        st.session_state.mask_settings,
    )
    st.session_state.cached_metrics = metrics
    st.session_state.measurements = metrics["measurements"]
    st.session_state.cached_overlay = overlay_calc
    st.session_state.cached_exports = {
        "json": safe_json(project),
        "csv": build_csv_from_metrics(metrics),
        "overlay_png": png_bytes(overlay_calc),
        "project": project,
    }
    st.session_state.metrics_dirty = False


# ════════════════════════════════════════════════════════════════════════
#  UI: шапка
# ════════════════════════════════════════════════════════════════════════
st.title("📐 SenseOptics Метролог")
st.markdown(
    f"<div class='so-version'>Версия: {APP_VERSION} · база: app334.py · пакетный ROI</div>",
    unsafe_allow_html=True,
)

fullscreen = st.toggle("🔍 Полноразмерный режим замеров", key="fullscreen_mode")

# ════════════════════════════════════════════════════════════════════════
#  UI: загрузка (скрыта в полноэкранном режиме)
# ════════════════════════════════════════════════════════════════════════
main_up = None
use_sample = False
sample = Path(__file__).parent / "sample" / "demo_template.png"

if not fullscreen:
    st.caption("Лёгкий режим для Replit/free: накопление ROI, расчёт по кнопке, экспорт и маска.")
    with st.container(border=True):
        st.markdown("**1. Загрузка изображения**")
        st.caption(
            "После успешной загрузки файл хранится в памяти сессии и в папке uploads на Replit. "
            "Если мобильный браузер потеряет upload при перерисовке, изображение восстановится автоматически."
        )
        main_up = st.file_uploader(
            "Перетащите файл сюда или нажмите Browse files",
            type=["jpg", "jpeg", "png", "bmp", "tif", "tiff", "webp"],
            key="main_upload",
            help="Загрузка находится в основной области. Боковая панель не используется.",
        )

        c_last, c_clear = st.columns(2)
        if c_last.button("🔁 Загрузить последнее сохранённое"):
            data_disk, name_disk, fid_disk, path_disk = load_last_image_from_disk()
            if data_disk and name_disk and fid_disk:
                st.session_state.stored_image_bytes = data_disk
                st.session_state.stored_image_name = name_disk
                st.session_state.stored_image_id = fid_disk
                st.session_state.stored_image_source = f"disk:{path_disk}"
                st.session_state.last_uploaded_path = path_disk
                st.rerun()
            else:
                st.warning("Последнее сохранённое изображение не найдено.")
        if c_clear.button("🧹 Забыть изображение"):
            st.session_state.stored_image_bytes = None
            st.session_state.stored_image_name = None
            st.session_state.stored_image_id = None
            st.session_state.stored_image_source = None
            clear_work_state(reset_calibration=True)
            st.rerun()

        if main_up is None and st.session_state.stored_image_bytes is None and sample.exists():
            use_sample = st.checkbox("Использовать demo_template.png", value=True)

# ════════════════════════════════════════════════════════════════════════
#  Получение байтов изображения (надёжно для мобильного/Replit)
# ════════════════════════════════════════════════════════════════════════
data: Optional[bytes] = None
src_name: Optional[str] = None
file_id: Optional[str] = None
load_note: Optional[str] = None

if main_up is not None:
    try:
        data = main_up.getvalue()
    except Exception:
        data = b""
    if data:
        src_name = main_up.name
        file_id = remember_image(data, src_name, source="upload")
        load_note = "из upload + сохранено в памяти"
    elif st.session_state.stored_image_bytes:
        data = st.session_state.stored_image_bytes
        src_name = st.session_state.stored_image_name or "stored_image"
        file_id = st.session_state.stored_image_id or file_fingerprint(data, src_name)
        load_note = "upload временно пустой, использована копия из памяти"
    else:
        st.warning("Файл выбран, но браузер не передал данные. Нажмите Browse files ещё раз или используйте последнее сохранённое.")
        st.stop()
elif st.session_state.stored_image_bytes:
    data = st.session_state.stored_image_bytes
    src_name = st.session_state.stored_image_name or "stored_image"
    file_id = st.session_state.stored_image_id or file_fingerprint(data, src_name)
    load_note = "из памяти сессии"
else:
    data_disk, name_disk, fid_disk, path_disk = load_last_image_from_disk()
    if data_disk and name_disk and fid_disk:
        data = data_disk
        src_name = name_disk
        file_id = fid_disk
        st.session_state.stored_image_bytes = data_disk
        st.session_state.stored_image_name = name_disk
        st.session_state.stored_image_id = fid_disk
        st.session_state.stored_image_source = f"disk:{path_disk}"
        st.session_state.last_uploaded_path = path_disk
        load_note = "автовосстановлено из uploads"
    elif use_sample and sample.exists():
        data = sample.read_bytes()
        src_name = sample.name
        file_id = remember_image(data, src_name, source="sample")
        load_note = "demo_template"
    else:
        if fullscreen:
            st.info("Сначала выключите полноразмерный режим и загрузите изображение.")
        else:
            st.info("Загрузите изображение. На телефоне и в Replit надёжнее пользоваться центральной областью загрузки.")
        st.stop()

img_bgr, err = decode_image(data, src_name)
if err or img_bgr is None:
    if st.session_state.stored_image_bytes:
        data = st.session_state.stored_image_bytes
        src_name = st.session_state.stored_image_name or "stored_image"
        file_id = st.session_state.stored_image_id or file_fingerprint(data, src_name)
        img_bgr, err = decode_image(data, src_name)
    if err or img_bgr is None:
        st.error(err or "Не удалось прочитать изображение.")
        st.stop()

if load_note and not fullscreen:
    st.caption(f"Источник изображения: {load_note} · размер файла: {len(data) / 1024:.1f} КБ")

orig_h, orig_w = img_bgr.shape[:2]

# ════════════════════════════════════════════════════════════════════════
#  UI: настройки и режим
# ════════════════════════════════════════════════════════════════════════
w_max = max(280, min(2400, int(orig_w)))
default_w = min(360, w_max)

if fullscreen:
    # Полноэкранный режим — только выбор режима (минимум элементов)
    mode = st.radio("Режим", MODES, horizontal=False, key="mode_select")
    max_w = int(st.session_state.get("work_width_value", default_w))
    max_w = min(max(280, max_w), w_max)
else:
    with st.container(border=True):
        st.markdown("**2. Настройки и режим**")
        max_w = st.slider(
            "Рабочая ширина, px",
            min_value=280,
            max_value=w_max,
            value=min(int(st.session_state.get("work_width_value", default_w)), w_max),
            step=20,
            key="work_width_value",
            help="При изменении ширины ROI сбрасываются, чтобы масштаб не сломался. На телефоне обычно лучше 280–420 px.",
        )

        st.markdown("**Калибровка (px/mm для рабочей картинки)**")
        ppm_value = float(st.session_state.px_per_mm or 0.0)
        cset1, cset2 = st.columns([2, 1])
        ppm_in = cset1.number_input(
            "px/mm вручную",
            min_value=0.0,
            value=ppm_value,
            step=0.5,
            format="%.4f",
            help="Можно ввести вручную или откалибровать в режиме «Калибровка».",
        )
        with cset2:
            if st.button("Применить px/mm") and ppm_in > 0:
                st.session_state.px_per_mm = float(ppm_in)
                st.session_state.calibration_info = {"source": "manual", "px_per_mm": float(ppm_in)}
                st.session_state.metrics_dirty = True
                st.rerun()
            if st.button("Сброс px/mm"):
                st.session_state.px_per_mm = None
                st.session_state.calibration_info = None
                st.session_state.metrics_dirty = True
                st.rerun()

        mode = st.radio("Режим", MODES, horizontal=False, key="mode_select")

        if mode == MODE_INCL:
            st.session_state.inclusion_fast_mode = st.checkbox(
                "⚡ Быстрый режим включений: клик = точка включения",
                value=bool(st.session_state.inclusion_fast_mode),
                help="Включено: каждый клик сразу создаёт ROI-точку. Выключено: точки копятся, затем «Добавить» создаёт их пачкой.",
            )
        if mode == MODE_MASK:
            shape_label = st.radio(
                "Форма ROI маски",
                ["Полигон", "Прямоугольник (2 клика)"],
                horizontal=True,
            )
            st.session_state.mask_roi_shape = "rect" if shape_label.startswith("Прямоуг") else "polygon"

# ════════════════════════════════════════════════════════════════════════
#  Подготовка рабочего изображения + контроль смены изображения
# ════════════════════════════════════════════════════════════════════════
work, scale = cached_downscale(img_bgr, int(max_w))
H, W = work.shape[:2]

current_image_id = f"{file_id}:{W}x{H}"
if st.session_state.current_image_id != current_image_id:
    clear_work_state(reset_calibration=True)
    st.session_state.current_image_id = current_image_id
    st.rerun()

ppm = st.session_state.px_per_mm

# ════════════════════════════════════════════════════════════════════════
#  Подсказки и статус
# ════════════════════════════════════════════════════════════════════════
hints = {
    MODE_LINE: "2 точки → «➕ Добавить линию». Длина считается только по «Посчитать метрики».",
    MODE_CIRCLE: "3 точки на окружности → «➕ Добавить окружность». Радиус/диаметр — по расчёту.",
    MODE_POLY: "Кликайте вершины (≥3) → «➕ Добавить полигон». Площадь — по расчёту.",
    MODE_INCL: "Кликайте включения. В быстром режиме каждый клик = точка. Подсчёт — по расчёту.",
    MODE_MASK: "Очертите область анализа → «➕ Добавить ROI маски». Маска строится отдельной кнопкой.",
    MODE_CALIB: "2 точки по эталону, длина и единицы ниже → «Применить калибровку».",
}
if not fullscreen:
    st.markdown(f"<div class='so-hint'>👆 {hints.get(mode, '')}</div>", unsafe_allow_html=True)

if st.session_state.cached_metrics is None:
    metric_state = "не рассчитаны"
elif st.session_state.metrics_dirty:
    metric_state = "требуют пересчёта"
else:
    metric_state = "актуальны"

dirty_cls = " so-dirty" if (st.session_state.metrics_dirty and st.session_state.cached_metrics is not None) else ""
st.markdown(
    f"<div class='so-status{dirty_cls}'>"
    f"Черновые точки: <b>{len(st.session_state.draft_points)}</b> · "
    f"ROI накоплено: <b>{len(st.session_state.draft_rois)}</b> · "
    f"Метрики: <b>{metric_state}</b>"
    f"{' · калибровка: %.4f px/mm' % ppm if ppm else ' · калибровка: не задана'}"
    f"</div>",
    unsafe_allow_html=True,
)

if not fullscreen:
    st.markdown(
        f"<div class='so-small'>Файл: {src_name} · оригинал: {orig_w}×{orig_h}px · "
        f"рабочее: {W}×{H}px · scale={scale:.4f}</div>",
        unsafe_allow_html=True,
    )

# ════════════════════════════════════════════════════════════════════════
#  Холст / клики (общий для обоих режимов)
# ════════════════════════════════════════════════════════════════════════
# Подбираем диапазон ширины кликабельного полотна; в полноэкранном — крупнее.
cw_max = int(min(1600, max(420, W * (4 if fullscreen else 2))))
cur_cw = int(st.session_state.get("canvas_display_width", 420))
cur_cw = max(280, min(cur_cw, cw_max))
st.session_state.canvas_display_width = cur_cw

canvas_display_width = st.slider(
    "Ширина кликабельного полотна на экране, px",
    min_value=280,
    max_value=cw_max,
    step=20,
    key="canvas_display_width",
    help="Только экранный размер полотна. Расчёты остаются в рабочих пикселях. Если на телефоне холст пропал — уменьшите до 360–520 px.",
)

# Для живого полотна используем закэшированные метрики (если они актуальны) —
# это просто подстановка значений в подписи, без какого-либо пересчёта.
live_meas = None
if st.session_state.cached_metrics is not None and not st.session_state.metrics_dirty:
    live_meas = st.session_state.cached_metrics.get("measurements")

overlay = draw_workspace(
    work,
    st.session_state.draft_rois,
    st.session_state.draft_points,
    mode,
    live_meas,
)

debug = False
if not fullscreen:
    show_control_preview = st.checkbox(
        "Показать некликабельную контрольную картинку",
        value=bool(st.session_state.get("show_control_preview_v5", False)),
        key="show_control_preview_v5",
        help="Только проверочная картинка. Для замеров кликайте по большому полотну ниже.",
    )
else:
    show_control_preview = False

if HAS_CLICK:
    if show_control_preview:
        st.image(
            overlay,
            caption="Контрольная картинка: клики НЕ принимает. Для замера кликайте по полотну ниже.",
            use_container_width=True,
        )

    canvas_scale = max(1e-9, float(canvas_display_width) / float(W))
    canvas_h = max(1, int(round(H * canvas_scale)))
    if int(canvas_display_width) != W:
        canvas_overlay = cv2.resize(overlay, (int(canvas_display_width), canvas_h), interpolation=cv2.INTER_LINEAR)
    else:
        canvas_overlay = overlay

    st.caption("Кликабельное полотно для замеров. Клик только добавляет точку/ROI и не запускает расчёт.")
    val = st_imgcoords(
        PILImage.fromarray(canvas_overlay),
        width=int(canvas_display_width),
        key=f"canvas_v5_{current_image_id}_{mode}_{int(canvas_display_width)}",
    )
    if val is not None and "x" in val and "y" in val:
        raw_x = float(val["x"])
        raw_y = float(val["y"])
        x = int(round(raw_x / canvas_scale))
        y = int(round(raw_y / canvas_scale))
        x, y = clamp_click(x, y, W, H)
        event_time = val.get("timestamp", val.get("time", val.get("event_time", "")))
        click_id = f"{current_image_id}:{mode}:{x}:{y}:{event_time}:{int(canvas_display_width)}"
        if click_id != st.session_state.last_click_id:
            st.session_state.last_click_id = click_id
            handle_click(x, y, mode)  # НЕ считает метрики
            st.rerun()
else:
    st.warning(
        "Пакет streamlit-image-coordinates не установлен или не загрузился. "
        "Проверьте requirements.txt и перезапустите Replit. Пока можно вводить координаты вручную."
    )
    st.image(overlay, use_container_width=True)
    mc1, mc2, mc3 = st.columns(3)
    mx = mc1.number_input("x", 0, W - 1, W // 2)
    my = mc2.number_input("y", 0, H - 1, H // 2)
    if mc3.button("Добавить точку"):
        handle_click(int(mx), int(my), mode)
        st.rerun()

# ════════════════════════════════════════════════════════════════════════
#  Кнопки пакетного режима (общие)
# ════════════════════════════════════════════════════════════════════════
add_labels = {
    MODE_LINE: "➕ Добавить линию",
    MODE_CIRCLE: "➕ Добавить окружность",
    MODE_POLY: "➕ Добавить полигон",
    MODE_INCL: "➕ Добавить точки включений",
    MODE_MASK: "➕ Добавить ROI маски",
    MODE_CALIB: "➕ (калибровка — кнопкой ниже)",
}

ra, rb = st.columns(2)
if ra.button("↩️ Удалить последнюю точку"):
    if undo_last_draft_point():
        st.rerun()
if rb.button("🧹 Очистить текущие точки"):
    st.session_state.draft_points = []
    st.rerun()

rc, rd = st.columns(2)
if rc.button(add_labels.get(mode, "➕ Добавить ROI"), disabled=(mode == MODE_CALIB)):
    ok, msg = commit_draft(mode, st.session_state.mask_roi_shape)
    if ok:
        st.rerun()
    else:
        st.warning(msg)
if rd.button("🗑 Удалить последний ROI"):
    if delete_last_roi():
        st.rerun()

calc_clicked = st.button("✅ Посчитать метрики", type="primary")
if calc_clicked:
    run_calculation(work, mode)
    if fullscreen:
        st.success("Метрики посчитаны. Выйдите из полноразмерного режима, чтобы увидеть таблицы и экспорт.")
    else:
        st.success("Метрики посчитаны.")

if fullscreen:
    if st.button("⛶ Выйти из полноразмерного режима"):
        st.session_state.fullscreen_mode = False
        st.rerun()

# ════════════════════════════════════════════════════════════════════════
#  Калибровка по двум кликам
# ════════════════════════════════════════════════════════════════════════
if mode == MODE_CALIB and len(st.session_state.draft_points) == 2:
    d_px = dist(st.session_state.draft_points[0], st.session_state.draft_points[1])
    st.info(f"Расстояние между точками: **{d_px:.1f} px**")
    kc1, kc2, kc3 = st.columns([1, 1, 1])
    known = kc1.number_input("Длина эталона", min_value=0.0001, value=10.0, step=1.0)
    units = kc2.selectbox("Единицы", ["mm", "µm"])
    if kc3.button("Применить калибровку") and known > 0:
        known_mm = known if units == "mm" else known / 1000.0
        st.session_state.px_per_mm = d_px / known_mm
        st.session_state.calibration_info = {
            "source": "2-click",
            "known_length": known,
            "units": units,
            "known_mm": known_mm,
            "px_distance": round(d_px, 3),
            "px_per_mm": st.session_state.px_per_mm,
        }
        st.session_state.metrics_dirty = True  # ROI не удаляются, но метрики надо пересчитать
        st.session_state.draft_points = []
        st.rerun()

# ════════════════════════════════════════════════════════════════════════
#  Дальше — только обычный режим (таблицы, маска, экспорт)
# ════════════════════════════════════════════════════════════════════════
if fullscreen:
    st.caption("Полноразмерный режим: таблицы, маска и экспорт скрыты. Выйдите из режима, чтобы их увидеть.")
    st.stop()

if st.session_state.metrics_dirty and st.session_state.cached_metrics is not None:
    st.warning("ROI изменены — нужно пересчитать метрики кнопкой «✅ Посчитать метрики».")

# ── Накопленные ROI ──────────────────────────────────────────────────────
with st.expander(f"🧱 Накопленные ROI ({len(st.session_state.draft_rois)})", expanded=False):
    if st.session_state.draft_rois:
        for i, roi in enumerate(st.session_state.draft_rois):
            cc1, cc2 = st.columns([6, 1])
            cc1.write(f"{i + 1}. `{roi['id']}` · {roi['type']} · {roi.get('label', '')} · точек: {len(roi['points'])}")
            if cc2.button("✕", key=f"del_roi_{roi['id']}"):
                st.session_state.draft_rois.pop(i)
                st.session_state.metrics_dirty = True
                st.rerun()
    else:
        st.caption("ROI ещё не добавлены. Накопите фигуры и нажмите «Посчитать метрики».")

# ── Результаты расчёта ───────────────────────────────────────────────────
metrics = st.session_state.cached_metrics
with st.expander(f"📏 Метрики ({len(st.session_state.measurements)})", expanded=True):
    if metrics is None:
        st.caption("Метрики ещё не рассчитаны. Нажмите «✅ Посчитать метрики».")
    else:
        for w in metrics.get("warnings", []):
            st.warning(w)
        for m in metrics["measurements"]:
            t = m["type"]
            if t == "line":
                txt = f"`{m['id']}` Линия: {m.get('length_px')} px" + (
                    f" = {m.get('length_mm')} мм" if m.get("length_mm") is not None else ""
                )
            elif t == "circle_3pt":
                txt = f"`{m['id']}` Окружность: D={m.get('diameter_px')} px, S={m.get('area_px2')} px²" + (
                    f" · D={m.get('diameter_mm')} мм, S={m.get('area_mm2')} мм²" if m.get("diameter_mm") is not None else ""
                )
            elif t == "polygon":
                txt = f"`{m['id']}` Полигон: {m.get('area_px2')} px² ({m.get('n_vertices')} верш.)" + (
                    f" = {m.get('area_mm2')} мм²" if m.get("area_mm2") is not None else ""
                )
            elif t == "inclusion_point":
                txt = f"`{m['id']}` Включение: ({m.get('x')}, {m.get('y')})"
            elif t == "mask_roi":
                txt = f"`{m['id']}` ROI маски: {m.get('roi_area_px2')} px²"
                if m.get("mask_percent_inside") is not None:
                    txt += f" · маска внутри: {m.get('mask_percent_inside')}%"
            else:
                txt = f"`{m['id']}` {t}"
            st.write(txt)

        st.markdown("**Сводка**")
        st.json(metrics.get("summary", {}))
        if metrics.get("statistics"):
            st.markdown("**Статистика**")
            st.json(metrics["statistics"])

        if st.session_state.cached_overlay is not None:
            st.image(st.session_state.cached_overlay, caption="Overlay после расчёта", use_container_width=True)

# ── Маска включений (ленивая) ────────────────────────────────────────────
with st.expander("🧩 Маска включений (строится только по кнопке)", expanded=False):
    mask_rois = [r for r in st.session_state.draft_rois if r.get("type") == "mask_roi"]
    if mask_rois:
        st.caption(f"Найдено ROI маски: {len(mask_rois)} — маска будет считаться только внутри них.")
    else:
        st.caption("ROI маски нет — маска будет считаться по всему рабочему изображению.")

    m1, m2, m3 = st.columns(3)
    threshold_mode = m1.radio("Порог", ["Otsu", "Manual"], horizontal=True)
    manual_threshold = m2.slider("Manual threshold", 0, 255, 90, 1)
    detect_bright = m3.checkbox("Искать светлые, не тёмные", value=False)

    m4, m5, m6, m7 = st.columns(4)
    min_area = m4.number_input("min area, px²", min_value=1, max_value=1000000, value=4, step=1)
    max_area = m5.number_input("max area, px² (0 = без лимита)", min_value=0, max_value=10000000, value=0, step=10)
    blur_size = m6.slider("blur", 1, 15, 3, 2)
    close_iter = m7.slider("close", 0, 5, 1, 1)

    mb1, mb2 = st.columns(2)
    if mb1.button("🧩 Построить / обновить маску"):
        region = _mask_roi_region(work, mask_rois)
        mask = build_inclusion_mask(
            work,
            threshold_mode=threshold_mode,
            manual_threshold=int(manual_threshold),
            min_area=int(min_area),
            max_area=int(max_area),
            detect_bright=bool(detect_bright),
            blur_size=int(blur_size),
            close_iter=int(close_iter),
            region=region,
        )
        st.session_state.mask = mask
        st.session_state.mask_overlay = make_mask_overlay(work, mask)
        st.session_state.mask_stats = mask_stats(mask, ppm)
        st.session_state.mask_settings = {
            "threshold_mode": threshold_mode,
            "manual_threshold": int(manual_threshold),
            "detect_bright": bool(detect_bright),
            "min_area_px": int(min_area),
            "max_area_px": int(max_area),
            "blur_size": int(blur_size),
            "close_iter": int(close_iter),
            "restricted_to_mask_roi": region is not None,
        }
        # Маска влияет на метрики mask_roi → пометим как dirty
        if mask_rois:
            st.session_state.metrics_dirty = True
        st.rerun()
    if mb2.button("🗑 Сбросить маску"):
        reset_mask()
        st.rerun()

    if st.session_state.mask is not None:
        st.image(st.session_state.mask_overlay, caption="Overlay маски", use_container_width=True)
        st.json(st.session_state.mask_stats)
        md1, md2 = st.columns(2)
        md1.download_button(
            "⬇️ Скачать mask.png",
            gray_png_bytes(st.session_state.mask),
            file_name=f"mask_{safe_stem(src_name)}.png",
            mime="image/png",
        )
        md2.download_button(
            "⬇️ Скачать mask_overlay.png",
            png_bytes(st.session_state.mask_overlay),
            file_name=f"mask_overlay_{safe_stem(src_name)}.png",
            mime="image/png",
        )
    else:
        st.caption("Маска ещё не построена.")

# ── Экспорт (после расчёта / по кнопке) ──────────────────────────────────
st.divider()
with st.expander("⬇️ Экспорт", expanded=True):
    if st.session_state.metrics_dirty and st.session_state.cached_metrics is not None:
        st.warning("ROI изменены — пересчитайте метрики перед экспортом, иначе экспорт устарел.")

    exports = st.session_state.cached_exports
    if exports is None:
        st.info("Экспорт станет доступен после нажатия «✅ Посчитать метрики».")
    else:
        e1, e2, e3 = st.columns(3)
        e1.download_button(
            "🖼 Overlay PNG",
            exports["overlay_png"],
            file_name=f"overlay_{safe_stem(src_name)}.png",
            mime="image/png",
        )
        e2.download_button(
            "⬇️ JSON проекта",
            exports["json"],
            file_name=f"project_{safe_stem(src_name)}.json",
            mime="application/json",
        )
        # Excel строим лениво по кнопке (openpyxl импортируется внутри функции).
        # CSV всегда доступен (готовится при расчёте). Чтобы не было
        # DuplicateWidgetID, кнопка measurements.csv создаётся ровно один раз.
        if HAS_XLSX:
            with e3:
                if st.button("📊 Собрать Excel"):
                    xlsx_bytes = build_xlsx_from_metrics(
                        st.session_state.cached_metrics, exports["project"]
                    )
                    st.session_state.cached_exports["xlsx"] = xlsx_bytes
                    st.rerun()
                if exports.get("xlsx"):
                    st.download_button(
                        "⬇️ measurements.xlsx",
                        exports["xlsx"],
                        file_name=f"measurements_{safe_stem(src_name)}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            # Excel занял колонку e3 — CSV выносим отдельной кнопкой ниже.
            st.download_button(
                "⬇️ measurements.csv",
                exports["csv"],
                file_name=f"measurements_{safe_stem(src_name)}.csv",
                mime="text/csv",
            )
        else:
            # openpyxl недоступен — в e3 сразу даём CSV (единственная CSV-кнопка).
            e3.download_button(
                "⬇️ measurements.csv",
                exports["csv"],
                file_name=f"measurements_{safe_stem(src_name)}.csv",
                mime="text/csv",
            )

        st.markdown("---")
        st.caption("Проект целиком (изображение как файл, без base64).")
        if st.button("💾 Собрать проект ZIP"):
            zip_bytes = build_project_zip(
                data,
                src_name,
                exports["project"],
                st.session_state.cached_metrics,
                st.session_state.cached_overlay,
                st.session_state.mask,
                st.session_state.mask_overlay,
            )
            st.session_state.cached_exports["zip"] = zip_bytes
            st.rerun()
        if exports.get("zip"):
            st.download_button(
                "⬇️ Скачать проект ZIP",
                exports["zip"],
                file_name=f"metrology_project_{safe_stem(src_name)}.zip",
                mime="application/zip",
            )
